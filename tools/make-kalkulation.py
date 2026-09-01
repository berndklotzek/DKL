# -*- coding: utf-8 -*-
"""Erzeugt die Margenkalkulation als Excel-Arbeitsmappe.

    python3 tools/make-kalkulation.py

Ergebnis: kalkulation/kaltstart-margenkalkulation.xlsx

Die Mappe enthält nur Formeln, keine vorberechneten Werte — Excel und
LibreOffice rechnen beim Öffnen. Verwendet werden ausschliesslich SUM, IF und
INDEX, damit die Datei in jeder Tabellenkalkulation funktioniert.

Die Erläuterung zu den Zahlen steht in docs/klimashop-kalkulation.md.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

ARIAL   = "Arial"
BLAU    = Font(name=ARIAL, size=10, color="0000FF")          # Eingaben
BLAU_B  = Font(name=ARIAL, size=10, color="0000FF", bold=True)
SCHWARZ = Font(name=ARIAL, size=10)
SCHWARZ_B = Font(name=ARIAL, size=10, bold=True)
GRUEN   = Font(name=ARIAL, size=10, color="008000")          # Verweis auf anderes Blatt
TITEL   = Font(name=ARIAL, size=14, bold=True, color="0C1A26")
UNTER   = Font(name=ARIAL, size=10, italic=True, color="5D6D7A")
KOPF    = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
GELB    = PatternFill("solid", fgColor="FFFF00")             # bitte prüfen
KOPFFILL= PatternFill("solid", fgColor="0B6FA8")
GRAU    = PatternFill("solid", fgColor="F4F8FB")
GRUEN_F = PatternFill("solid", fgColor="E3F4ED")
LINIE   = Border(bottom=Side(style="thin", color="C9D8E6"))
DICK    = Border(top=Side(style="medium", color="0B6FA8"))

EUR  = '#,##0.00 "€";(#,##0.00 "€");-'
EUR0 = '#,##0 "€";(#,##0 "€");-'
USD  = '#,##0.00 "$";(#,##0.00 "$");-'
PROZ = '0.0%'
STK  = '#,##0'
FAKT = '0.00"×"'

wb = Workbook()

# ===========================================================================
# 1 Anleitung
# ===========================================================================
ws = wb.active
ws.title = "Anleitung"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 104

zeilen = [
    ("t", "KALTSTART Klimasysteme — Margenkalkulation"),
    ("u", "Erstellt am 31.08.2026. Alle Zahlen sind Annahmen und müssen durch echte Angebote ersetzt werden."),
    ("", ""),
    ("h", "So benutzen Sie diese Mappe"),
    ("", "Alles, was BLAU geschrieben ist, dürfen und sollen Sie ändern. Schwarze Zellen sind Formeln —"),
    ("", "wenn Sie die überschreiben, rechnet die Mappe falsch."),
    ("", "GELB hinterlegte Zellen sind Annahmen, die Sie zuerst prüfen sollten. Vor allem der Wechselkurs."),
    ("", ""),
    ("h", "Die vier Blätter"),
    ("", "Annahmen      Alles, was für alle Geräte gleich gilt: Wechselkurs, Fracht, Zoll, Gebühren, Retouren."),
    ("", "Kalkulation   Einkaufspreis, Landepreis und Deckungsbeitrag je Gerät. Das Kernstück."),
    ("", "Container     Was eine Erstbestellung kostet und ab welcher Verkaufsquote sie sich lohnt."),
    ("", "Marktpreise   Was Wettbewerber in Deutschland verlangen. Recherchiert, mit Quelle."),
    ("", ""),
    ("h", "Die drei Begriffe, um die es geht"),
    ("", "Einkaufspreis (FOB)   Was die Fabrik verlangt, geliefert bis zum Hafen in China. Ohne Fracht,"),
    ("", "                      ohne Zoll, ohne alles Weitere."),
    ("", "Landepreis            Was das Gerät kostet, wenn es bei Ihnen im Lager steht. FOB plus Fracht,"),
    ("", "                      Zoll, Entsorgungsgebühren und anteilige Zertifizierungskosten."),
    ("", "                      Das ist Ihr echter Einkaufspreis — nicht der FOB-Preis."),
    ("", "Deckungsbeitrag       Was nach Abzug ALLER variablen Kosten übrig bleibt. Davon bezahlen Sie"),
    ("", "                      Lager, Werbung, Versicherung, Steuern und sich selbst."),
    ("", ""),
    ("h", "Der häufigste Anfängerfehler"),
    ("", "„Ich kaufe für 165 € ein und verkaufe für 649 €, also verdiene ich 484 €.“ Falsch."),
    ("", "Von den 649 € gehen zuerst 104 € Mehrwertsteuer ab. Dann Zahlungsgebühr, Versand, den Sie ab"),
    ("", "499 € geschenkt haben, eine Rückstellung für Retouren und eine für Gewährleistungsfälle."),
    ("", "Was übrig bleibt, steht auf dem Blatt „Kalkulation“ in der Zeile Deckungsbeitrag."),
    ("", ""),
    ("h", "Beim ersten Öffnen"),
    ("", "Excel und LibreOffice berechnen die Mappe beim Öffnen. Sollte eine Vorschau-App leere"),
    ("", "Zellen zeigen, öffnen Sie die Datei in Excel oder LibreOffice — dann stehen die Zahlen da."),
    ("", ""),
    ("h", "Woher die Zahlen stammen"),
    ("", "Verkaufspreise der Wettbewerber und FOB-Spannen sind recherchiert (Blatt „Marktpreise“)."),
    ("", "Fracht, Zoll, Gebühren und Quoten sind branchenübliche Erfahrungswerte, keine Angebote."),
    ("", "Sobald das erste Angebot aus China vorliegt, tragen Sie den echten FOB-Preis ein."),
]
r = 2
for art, text in zeilen:
    c = ws.cell(row=r, column=2, value=text)
    if art == "t":   c.font = TITEL
    elif art == "u": c.font = UNTER
    elif art == "h":
        c.font = Font(name=ARIAL, size=11, bold=True, color="0B6FA8")
    else:            c.font = SCHWARZ
    r += 1

# ===========================================================================
# 2 Annahmen
# ===========================================================================
an = wb.create_sheet("Annahmen")
an.sheet_view.showGridLines = False
an.column_dimensions["A"].width = 3
an.column_dimensions["B"].width = 46
an.column_dimensions["C"].width = 14
an.column_dimensions["D"].width = 58

an["B2"] = "Annahmen"; an["B2"].font = TITEL
an["B3"] = "Blau = ändern Sie hier. Gelb = zuerst prüfen."; an["B3"].font = UNTER

def kopfzeile(ws, zeile, titel):
    for sp, txt in zip("BCD", [titel, "Wert", "Erläuterung"]):
        c = ws[sp + str(zeile)]; c.value = txt; c.font = KOPF; c.fill = KOPFFILL
        c.alignment = Alignment(horizontal="left")

def eingabe(ws, zeile, bez, wert, fmt, hinweis, gelb=False):
    ws["B%d" % zeile] = bez; ws["B%d" % zeile].font = SCHWARZ
    c = ws["C%d" % zeile]; c.value = wert; c.font = BLAU; c.number_format = fmt
    if gelb: c.fill = GELB
    ws["D%d" % zeile] = hinweis; ws["D%d" % zeile].font = UNTER
    for sp in "BCD": ws[sp + str(zeile)].border = LINIE

kopfzeile(an, 5, "Einkauf und Import")
eingabe(an, 6,  "Wechselkurs (EUR je USD)", 0.92, '0.000',
        "BITTE PRÜFEN — Tageskurs eintragen. Ändert alle Einkaufspreise.", gelb=True)
eingabe(an, 7,  "Seefracht je 40-Fuß-Container (EUR)", 3200, EUR0,
        "China–Nordeuropa. Schwankt stark, 2.500 bis 4.500 € sind üblich.")
eingabe(an, 8,  "Zollsatz auf Klimageräte", 0.022, PROZ,
        "Position 8415. Berechnungsgrundlage ist Warenwert plus Fracht.")
eingabe(an, 9,  "Entsorgungs- und Verpackungsgebühren je Gerät (EUR)", 3.00, EUR,
        "WEEE-Beteiligung, duales System, Batteriegesetz.")
eingabe(an, 10, "Einmalkosten je Modell (EUR)", 6000, EUR0,
        "Prüfberichte, EPREL, Übersetzung der Anleitung, Fabrikaudit, Muster.")

kopfzeile(an, 12, "Verkauf")
eingabe(an, 13, "Mehrwertsteuer", 0.19, PROZ, "Regelbesteuerung.")
eingabe(an, 14, "Zahlungsgebühr (vom Bruttoumsatz)", 0.020, PROZ,
        "Karte über Stripe ca. 1,5 %, PayPal und Klarna deutlich mehr. Mischsatz.")
eingabe(an, 15, "Versandkosten je Sendung, die Sie tragen (EUR)", 32.00, EUR,
        "Spedition innerhalb Deutschlands. Fällt an, wenn Sie versandkostenfrei liefern.")
eingabe(an, 16, "Schwelle für versandkostenfreie Lieferung (EUR)", 499, EUR0,
        "Steht so auf der Website und in stripe/catalog.json.")

kopfzeile(an, 18, "Risiken")
eingabe(an, 19, "Retourenquote", 0.05, PROZ,
        "Fernabsatz mit 14 Tagen Widerruf. 3 bis 6 % sind bei Weißware normal.")
eingabe(an, 20, "Wertverlust je Retoure (vom Landepreis)", 0.40, PROZ,
        "Ein zurückgesendetes Gerät ist oft nur noch als B-Ware verkäuflich.")
eingabe(an, 21, "Rückstellung Gewährleistung (vom Nettoumsatz)", 0.025, PROZ,
        "Zwei Jahre Gewährleistung. Fälle kommen meist im zweiten Sommer.")

an["B23"] = "Hinweis"; an["B23"].font = SCHWARZ_B
an["B24"] = ("Die Einfuhrumsatzsteuer taucht hier bewusst nicht auf: Sie zahlen sie beim Import, "
             "holen sie aber als Vorsteuer zurück. Sie ist kein Kostenfaktor, sondern nur "
             "Liquidität — und steht deshalb auf dem Blatt „Container“.")
an["B24"].font = UNTER
an.merge_cells("B24:D26")
an["B24"].alignment = Alignment(wrap_text=True, vertical="top")
an["C6"].comment = Comment("Kurs vom Tag der Bestellung eintragen. Bei 300 Geräten "
                           "verschiebt ein Kursunterschied von 5 % rund 2.500 € Einkaufswert.",
                           "Kalkulation", height=110, width=280)

# ===========================================================================
# 3 Kalkulation
# ===========================================================================
ka = wb.create_sheet("Kalkulation")
ka.sheet_view.showGridLines = False
ka.column_dimensions["A"].width = 3
ka.column_dimensions["B"].width = 44
for sp in "CDEF": ka.column_dimensions[sp].width = 16
ka.column_dimensions["G"].width = 46

ka["B2"] = "Kalkulation je Gerät"; ka["B2"].font = TITEL
ka["B3"] = ("Links die Geräte ohne Montage, rechts die zur Selbstmontage mit Außengerät. "
            "Blaue Zellen ändern.")
ka["B3"].font = UNTER

PRODUKTE = [
    ("Zweikanal 9.000 BTU", 150, 320, 549, 449),
    ("Zweikanal 12.000 BTU", 180, 260, 649, 549),
    ("Quick-Connect 9.000 BTU", 175, 200, 649, 477),
    ("Quick-Connect 12.000 BTU", 200, 170, 749, 570),
]

for sp, (name, *_ ) in zip("CDEF", PRODUKTE):
    c = ka[sp + "5"]; c.value = name; c.font = KOPF; c.fill = KOPFFILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)
ka["B5"] = "Ohne Montage  ·  Selbstmontage"; ka["B5"].font = KOPF; ka["B5"].fill = KOPFFILL
ka["G5"] = "Erläuterung"; ka["G5"].font = KOPF; ka["G5"].fill = KOPFFILL
ka.row_dimensions[5].height = 30

def zeile(nr, bez, formeln, fmt, hinweis, fett=False, fill=None, font=None):
    ka["B%d" % nr] = bez
    ka["B%d" % nr].font = SCHWARZ_B if fett else SCHWARZ
    for sp, wert in zip("CDEF", formeln):
        c = ka[sp + str(nr)]; c.value = wert
        c.font = font or (SCHWARZ_B if fett else SCHWARZ)
        c.number_format = fmt
        if fill: c.fill = fill
    ka["G%d" % nr] = hinweis; ka["G%d" % nr].font = UNTER
    for sp in "BCDEFG": ka[sp + str(nr)].border = LINIE
    if fill: ka["B%d" % nr].fill = fill

# --- Einkauf
ka["B7"] = "EINKAUF"; ka["B7"].font = Font(name=ARIAL, size=10, bold=True, color="0B6FA8")
zeile(8, "Einkaufspreis FOB je Gerät (USD)", [p[1] for p in PRODUKTE], USD,
      "Angebot der Fabrik, geliefert bis Hafen China. Hier den echten Wert eintragen.",
      font=BLAU)
zeile(9, "Stück je 40-Fuß-Container", [p[2] for p in PRODUKTE], STK,
      "Fragen Sie den Lieferanten. Split-Sets brauchen mehr Platz als Mobilgeräte.",
      font=BLAU)
zeile(10, "Einkaufspreis FOB je Gerät (EUR)",
      ["=%s8*Annahmen!$C$6" % sp for sp in "CDEF"], EUR,
      "FOB in Euro, zum Kurs aus dem Blatt Annahmen.")
zeile(11, "Seefracht je Gerät",
      ["=Annahmen!$C$7/%s9" % sp for sp in "CDEF"], EUR,
      "Containerfracht geteilt durch die Stückzahl.")
zeile(12, "Zoll je Gerät",
      ["=(%s10+%s11)*Annahmen!$C$8" % (sp, sp) for sp in "CDEF"], EUR,
      "Auf Warenwert plus Fracht.")
zeile(13, "Entsorgung und Verpackung je Gerät",
      ["=Annahmen!$C$9" for _ in "CDEF"], EUR,
      "WEEE, duales System, Batteriegesetz.")
zeile(14, "Einmalkosten je Gerät",
      ["=Annahmen!$C$10/%s9" % sp for sp in "CDEF"], EUR,
      "Zertifizierung und Muster, verteilt auf die erste Containerladung.")
zeile(15, "Landepreis je Gerät", ["=SUM(%s10:%s14)" % (sp, sp) for sp in "CDEF"], EUR,
      "Ihr echter Einkaufspreis, wenn das Gerät im Lager steht.",
      fett=True, fill=GRAU)

# --- Verkauf
ka["B17"] = "VERKAUF"; ka["B17"].font = Font(name=ARIAL, size=10, bold=True, color="0B6FA8")
zeile(18, "Ihr Verkaufspreis (brutto)", [p[3] for p in PRODUKTE], EUR0,
      "Preis auf der Website, inklusive Mehrwertsteuer.", font=BLAU_B)
zeile(19, "Vergleichspreis Wettbewerb (brutto)", [p[4] for p in PRODUKTE], EUR0,
      "Recherchierte Marktpreise, siehe Blatt Marktpreise.", font=BLAU)
zeile(20, "Abstand zum Wettbewerb",
      ["=IF(%s19=0,0,%s18/%s19-1)" % (sp, sp, sp) for sp in "CDEF"], PROZ,
      "Positiv heißt: Sie sind teurer. Das muss ein Argument rechtfertigen.")
zeile(21, "Nettoumsatz je Gerät",
      ["=%s18/(1+Annahmen!$C$13)" % sp for sp in "CDEF"], EUR,
      "Was nach Abzug der Mehrwertsteuer bei Ihnen ankommt.")

# --- Abzüge
ka["B23"] = "WAS VOM VERKAUFSPREIS ABGEHT"
ka["B23"].font = Font(name=ARIAL, size=10, bold=True, color="0B6FA8")
zeile(24, "Landepreis (Wareneinsatz)", ["=%s15" % sp for sp in "CDEF"], EUR,
      "Aus dem Abschnitt Einkauf.", font=GRUEN)
zeile(25, "Zahlungsgebühr",
      ["=%s18*Annahmen!$C$14" % sp for sp in "CDEF"], EUR,
      "Stripe, PayPal, Klarna. Wird auf den Bruttobetrag berechnet.")
zeile(26, "Versandkosten, die Sie tragen",
      ["=IF(%s18>=Annahmen!$C$16,Annahmen!$C$15,0)" % sp for sp in "CDEF"], EUR,
      "Ab der Freigrenze zahlen Sie den Versand. Das ist echtes Geld.")
zeile(27, "Retourenrisiko",
      ["=Annahmen!$C$19*(%s15*Annahmen!$C$20+Annahmen!$C$15*2)" % sp for sp in "CDEF"], EUR,
      "Quote × (Wertverlust am Gerät + Hin- und Rückversand).")
zeile(28, "Rückstellung Gewährleistung",
      ["=%s21*Annahmen!$C$21" % sp for sp in "CDEF"], EUR,
      "Für Reparaturen und Austauschgeräte im zweiten Jahr.")

# --- Ergebnis
zeile(30, "Deckungsbeitrag je Gerät",
      ["=%s21-%s24-%s25-%s26-%s27-%s28" % ((sp,)*6) for sp in "CDEF"], EUR,
      "Was übrig bleibt. Davon zahlen Sie Lager, Werbung, Versicherung und sich selbst.",
      fett=True, fill=GRUEN_F)
zeile(31, "Marge (vom Nettoumsatz)",
      ["=IF(%s21=0,0,%s30/%s21)" % (sp, sp, sp) for sp in "CDEF"], PROZ,
      "Unter 30 % wird es im Onlinehandel eng, über 45 % ist gut.", fett=True)
zeile(32, "Aufschlagsfaktor auf den Landepreis",
      ["=IF(%s15=0,0,%s18/%s15)" % (sp, sp, sp) for sp in "CDEF"], FAKT,
      "Im Elektrohandel sind 2,5× bis 3,5× auf den Landepreis üblich.")
zeile(33, "Niedrigster Verkaufspreis ohne Verlust (brutto)",
      ["=(%s24+%s26+%s27)*(1+Annahmen!$C$13)/"
       "((1-Annahmen!$C$21)-Annahmen!$C$14*(1+Annahmen!$C$13))" % (sp, sp, sp) for sp in "CDEF"],
      EUR0,
      "Darunter legen Sie bei jedem Gerät drauf. Versandkosten wie beim aktuellen Preis angesetzt.")

ka["B35"] = "Lesehilfe"; ka["B35"].font = SCHWARZ_B
ka["B36"] = ("Der Aufschlagsfaktor ist die Zahl, die Händler im Kopf haben. Verkaufspreis geteilt "
             "durch Landepreis. Unter 2,5 verdienen Sie im Onlinehandel nach Werbung und Retouren "
             "meist nichts mehr, weil Zahlungsgebühr, Versand und Rückstellungen den Rest fressen.")
ka["B36"].font = UNTER
ka.merge_cells("B36:G38")
ka["B36"].alignment = Alignment(wrap_text=True, vertical="top")

# ===========================================================================
# 4 Container
# ===========================================================================
co = wb.create_sheet("Container")
co.sheet_view.showGridLines = False
co.column_dimensions["A"].width = 3
co.column_dimensions["B"].width = 46
co.column_dimensions["C"].width = 16
co.column_dimensions["D"].width = 58

co["B2"] = "Was eine Erstbestellung kostet"; co["B2"].font = TITEL
co["B3"] = "Beispiel für ein Modell. Wählen Sie über die blaue Zelle, welches."
co["B3"].font = UNTER

kopfzeile(co, 5, "Bestellung")
co["B6"] = "Modell (Spalte C, D, E oder F der Kalkulation)"; co["B6"].font = SCHWARZ
co["B6"] = "Welches Modell? (1 bis 4)"
co["C6"] = 2; co["C6"].font = BLAU; co["C6"].fill = GELB; co["C6"].number_format = "0"
co["D6"] = "1 = Zweikanal 9k · 2 = Zweikanal 12k · 3 = Quick-Connect 9k · 4 = Quick-Connect 12k"
co["D6"].font = UNTER

def rechnung(zeile, bez, formel, fmt, hinweis, fett=False, fill=None):
    co["B%d" % zeile] = bez
    co["B%d" % zeile].font = SCHWARZ_B if fett else SCHWARZ
    c = co["C%d" % zeile]; c.value = formel
    c.font = SCHWARZ_B if fett else SCHWARZ; c.number_format = fmt
    if fill: c.fill = fill; co["B%d" % zeile].fill = fill
    co["D%d" % zeile] = hinweis; co["D%d" % zeile].font = UNTER
    for sp in "BCD": co[sp + str(zeile)].border = LINIE

# Über INDEX/MATCH auf die gewählte Spalte zugreifen
SP = '=INDEX(Kalkulation!$C${r}:$F${r},1,$C$6)'
rechnung(7, "Stück je Container", SP.format(r=9), STK, "Aus der Kalkulation.")
rechnung(8, "Landepreis je Gerät", SP.format(r=15), EUR, "Aus der Kalkulation.")
rechnung(9, "Verkaufspreis brutto", SP.format(r=18), EUR0, "Aus der Kalkulation.")
rechnung(10, "Deckungsbeitrag je Gerät", SP.format(r=30), EUR, "Aus der Kalkulation.")

kopfzeile(co, 12, "Kapitalbedarf")
rechnung(13, "Wareneinsatz (FOB gesamt)",
         "=INDEX(Kalkulation!$C$10:$F$10,1,$C$6)*C7", EUR0,
         "Das zahlen Sie der Fabrik, meist 30 % im Voraus.")
rechnung(14, "Seefracht", "=Annahmen!C7", EUR0, "Ein Container.")
rechnung(15, "Zoll", "=(C13+C14)*Annahmen!C8", EUR0, "Auf Warenwert plus Fracht.")
rechnung(16, "Einmalkosten Zertifizierung", "=Annahmen!C10", EUR0,
         "Fällt nur beim ersten Container je Modell an.")
rechnung(17, "Entsorgungs- und Verpackungsgebühren", "=Annahmen!C9*C7", EUR0, "Für die ganze Menge.")
rechnung(18, "Gebundenes Kapital", "=SUM(C13:C17)", EUR0,
         "So viel Geld liegt im Lager, bevor Sie den ersten Euro einnehmen.",
         fett=True, fill=GRAU)
rechnung(19, "Einfuhrumsatzsteuer (nur Liquidität)", "=(C13+C14+C15)*Annahmen!C13", EUR0,
         "Zahlen Sie beim Import, holen Sie als Vorsteuer zurück. Kein Kostenfaktor, "
         "aber Sie brauchen das Geld zwischenzeitlich.")

kopfzeile(co, 21, "Ergebnis, wenn alles verkauft wird")
rechnung(22, "Umsatz brutto", "=C9*C7", EUR0, "Bei 100 % Verkaufsquote.")
rechnung(23, "Deckungsbeitrag gesamt", "=C10*C7", EUR0,
         "Vor Lager, Werbung, Versicherung und Steuern.", fett=True, fill=GRUEN_F)

kopfzeile(co, 25, "Ab wann lohnt es sich")
rechnung(26, "Geldrückfluss je verkauftem Gerät", "=C10+C8", EUR,
         "Deckungsbeitrag plus Wareneinsatz — der Wareneinsatz steckt schon im "
         "gebundenen Kapital und darf hier nicht doppelt abgezogen werden.")
rechnung(27, "Stückzahl für Kostendeckung", "=IF(C26=0,0,C18/C26)", STK,
         "So viele Geräte müssen verkauft sein, bis das eingesetzte Geld wieder da ist.",
         fett=True)
rechnung(28, "Das entspricht einer Verkaufsquote von",
         "=IF(C7=0,0,C27/C7)", PROZ,
         "Bezogen auf die Containermenge.", fett=True, fill=GRUEN_F)
rechnung(29, "Ergebnis bei 70 % Verkaufsquote", "=C10*C7*0.7-Annahmen!C10*0.3", EUR0,
         "Deckungsbeitrag aus 70 % der Menge. Die Einmalkosten für die unverkauften "
         "30 % sind noch nicht wieder eingespielt und daher abgezogen. Der Wareneinsatz "
         "der Restmenge bleibt als Lagerwert erhalten.", fett=True)

co["B31"] = "Was diese Rechnung nicht enthält"; co["B31"].font = SCHWARZ_B
co["B32"] = ("Lagermiete, Werbung, Ihre Arbeitszeit, Versicherung, Steuerberatung, Einkommensteuer. "
             "Und den Fall, dass der Sommer 2027 kühl wird und die Hälfte des Containers ein Jahr "
             "länger steht. Rechnen Sie mit der 70-Prozent-Zeile, nicht mit der 100-Prozent-Zeile.")
co["B32"].font = UNTER
co.merge_cells("B32:D34")
co["B32"].alignment = Alignment(wrap_text=True, vertical="top")

# ===========================================================================
# 5 Marktpreise
# ===========================================================================
mk = wb.create_sheet("Marktpreise")
mk.sheet_view.showGridLines = False
mk.column_dimensions["A"].width = 3
mk.column_dimensions["B"].width = 40
mk.column_dimensions["C"].width = 15
mk.column_dimensions["D"].width = 20
mk.column_dimensions["E"].width = 52

mk["B2"] = "Was der Markt verlangt"; mk["B2"].font = TITEL
mk["B3"] = "Recherchiert am 31.08.2026. Preise ändern sich — vor der Preisfestsetzung nachprüfen."
mk["B3"].font = UNTER

for sp, txt in zip("BCDE", ["Produkt", "Preis", "Anbieter", "Quelle / Anmerkung"]):
    c = mk[sp + "5"]; c.value = txt; c.font = KOPF; c.fill = KOPFFILL

markt = [
    ("GERÄTE OHNE MONTAGE", None, None, None),
    ("Mobiles Klimagerät 12.000 BTU, bis 48 m²", 449, "MediaMarkt",
     "Einschlauchgerät, Standardware im Elektrohandel"),
    ("Testsieger-Klasse Stiftung Warentest", 400, "Handel allgemein",
     "Spanne der gut bewerteten Geräte: rund 300 bis 500 €"),
    ("De'Longhi Pinguino PAC EX Serie", 600, "Fachhandel",
     "Zweischlauch-Premium, je nach Modell 500 bis 800 €"),
    ("", None, None, None),
    ("SELBSTMONTAGE MIT AUSSENGERÄT", None, None, None),
    ("Heiko 12.000 BTU R32 Split", 477, "eBay",
     "inklusive Versand"),
    ("BeCool BC12SK2101QW Quick Connect", 570, "eBay",
     "zuzüglich 23,99 € Versand, A++, R32 vorgefüllt"),
    ("DANYON XA35QC Quick Connect 12.000 BTU", None, "danyon.de",
     "mit 5-m-Montageset; Anbieter positioniert es für den Fachbetrieb"),
    ("Weitere Anbieter im deutschen Markt", None, "Kältebringer, KlimaWorld, environ",
     "Die Nische ist bereits besetzt — siehe Anmerkung unten"),
]
r = 6
for bez, preis, anbieter, quelle in markt:
    if preis is None and anbieter is None:
        if bez:
            c = mk["B%d" % r]; c.value = bez
            c.font = Font(name=ARIAL, size=10, bold=True, color="0B6FA8")
        r += 1
        continue
    mk["B%d" % r] = bez; mk["B%d" % r].font = SCHWARZ
    if preis:
        c = mk["C%d" % r]; c.value = preis; c.font = SCHWARZ; c.number_format = EUR0
    else:
        mk["C%d" % r] = "k. A."; mk["C%d" % r].font = UNTER
    mk["D%d" % r] = anbieter; mk["D%d" % r].font = SCHWARZ
    mk["E%d" % r] = quelle; mk["E%d" % r].font = UNTER
    for sp in "BCDE": mk[sp + str(r)].border = LINIE
    r += 1

r += 1
mk["B%d" % r] = "Die wichtigste Erkenntnis aus dieser Recherche"
mk["B%d" % r].font = Font(name=ARIAL, size=11, bold=True, color="A8630B")
r += 1
mk["B%d" % r] = (
    "Die beiden Produktlinien stehen in völlig unterschiedlichen Märkten.\n\n"
    "Selbstmontage mit Außengerät: Hier gibt es in Deutschland bereits DANYON, BeCool, "
    "Kältebringer, KlimaWorld und environ, mit Preisen zwischen 477 und 570 € für 12.000 BTU. "
    "Der Preis auf Ihrer Website liegt darüber. In diesen Markt kommen Sie nur über den Preis "
    "hinein — und Preiskampf gegen eingeführte Anbieter ist als Neueinsteiger die schlechteste "
    "aller Ausgangslagen.\n\n"
    "Geräte ohne Montage: Der Markt ist groß, aber fast vollständig mit Einschlauchgeräten "
    "besetzt (MediaMarkt 449 € für 12.000 BTU). Zweischlauchgeräte gibt es fast nur von "
    "De'Longhi im Premiumbereich. Dazwischen ist Platz — und Sie haben ein technisches "
    "Argument, das stimmt und das der Kunde versteht.\n\n"
    "Empfehlung: Starten Sie mit einer Linie, nicht mit zweien. Die ohne Montage."
)
mk["B%d" % r].font = SCHWARZ
mk.merge_cells("B%d:E%d" % (r, r + 12))
mk["B%d" % r].alignment = Alignment(wrap_text=True, vertical="top")

wb.save("/home/user/DKL/kalkulation/kaltstart-margenkalkulation.xlsx")
print("Mappe geschrieben")
